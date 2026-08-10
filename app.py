import os
import json
from urllib import response
from flask import Flask, request, jsonify, send_file
from dotenv import load_dotenv
from openai import OpenAI
from flask_cors import CORS
from google_drive import (
    upload_file,
    download_file,
    update_file
)
import requests
import traceback
import uuid
from openpyxl.styles import PatternFill
from datetime import datetime
import gspread
from google.oauth2.service_account import Credentials


current_year = datetime.now().year
load_dotenv(override=True)

client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
key = os.getenv("OPENAI_API_KEY")
#print("Last 10:", key[-10:])

app = Flask(
    __name__,
    static_folder="frontend",
    static_url_path=""
)
CORS(app)

# ---------------------------
# In-memory DB
# ---------------------------
DB = []
LAST_UPLOADED_FILE = None
LAST_DRIVE_FILE_ID = None

def get_employee_sheet():

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive"
    ]

    creds = Credentials.from_service_account_file(
        "service-account.json",
        scopes=scopes
    )

    print("===== SERVICE ACCOUNT USED =====")
    print(creds.service_account_email)

    client = gspread.authorize(creds)

    spreadsheet = client.open_by_key(
        "1LX121_6X4ptcBb0DycWWn5yqCJd3AppMbTxGk78MqPc"
    )

    sheet = spreadsheet.worksheet("Sheet1")

    return sheet

# ---------------------------
# 4. Extract tasks
# ---------------------------
from openpyxl import load_workbook

def extract_tasks(file):

    wb = load_workbook(file)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]

    open_col = headers.index("Open")
    close_col = headers.index("Close")

    status_map = {
        "FF6AA84F": "Done",
        "FFB7B7B7": "Due",
        "FF999999": "Due",
        "FF6FA8DC": "Half-Done",
        "FFFF0000": "Redo",
        "FFE84499": "Late",
        "FFBF8E00": "On Hold",
        "FF00FFFF": "Almost Ready",
        "FF7A3F00": "Just Started",
        "FFFFF2CC": "NA"
    }

    tasks = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
        task = row[1].value
        print("Checking row", row_num, "Task:", task)
        
        employee_cells = [
            row[i].value
            for i in range(2, open_col)
            ]
        if all(v is None for v in employee_cells):
            continue

        # Skip blank rows
        if not task:
            continue
    
        
        open_date = row[open_col].value
        close_date = row[close_col].value
        
        for col_idx in range(2, open_col):
            employee = headers[col_idx]
            cell = row[col_idx]

            color = str(cell.fill.start_color.rgb)
            #print(f"Task={task}, Employee={employee}, Color={color}")

            print(task, employee, color)
            
            status = status_map.get(color)

            if status:
                tasks.append({
                    "task": task,
                    "employee": employee,
                    "status": status,
                    "color": "#" + color[-6:],   
                    "open": str(open_date),
                    "close": str(close_date)
                    })
                    
    print("Total tasks extracted:", len(tasks))
    
    if tasks:
        print("Last task:", tasks[-1])

        print("\n========== EXTRACTED TASKS ==========")
        for t in tasks[-10:]:
            print(t)
        print("Total extracted:", len(tasks))

        return tasks

N8N_WEBHOOK = "https://excelchatbot-n8n-production.up.railway.app/webhook/task-manager"

def send_to_n8n(
    action,
    task,
    employees,
    open_date,
    close_date,
    drive_file_id,
    field="",
    value="",
    updates=None
):
    try:

        emails = []

        for employee in employees:
            email = get_employee_email(employee)

            if email:
                emails.append(email)

        payload = {
            "action": action,
            "task": task,
            "employees": employees,
            "emails": emails,
            "open": open_date,
            "close": close_date,
            "field": field,
            "value": value,
            "updates": updates or [],
            "drive_file_id": drive_file_id
        }

        print("========== BEFORE N8N ==========")
        print("PAYLOAD:", payload)

        res = requests.post(
            N8N_WEBHOOK,
            json=payload,
            timeout=10
        )

        print("========== AFTER N8N ==========")
        print("STATUS:", res.status_code)
        print("BODY:", res.text)

        if 200 <= res.status_code < 300:
            print("N8N SUCCESS")
            return True

        print("N8N FAILED")
        return False

    except Exception as e:
        print("========== N8N REQUEST EXCEPTION ==========")
        print("ERROR:", repr(e))
        import traceback
        traceback.print_exc()
        return False
    
def find_task_details(excel_path, task_name):

    wb = load_workbook(excel_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    open_col = headers.index("Open")

    status_colors = {
        "FF6AA84F",
        "FFB7B7B7",
        "FF999999",
        "FF6FA8DC",
        "FFFF0000",
        "FFE84499",
        "FFBF8E00",
        "FF00FFFF",
        "FF7A3F00",
        "FFFFF2CC"
    }

    print("Looking for:", task_name)

    employees = []

    for r in range(2, ws.max_row + 1):

        excel_task = str(ws.cell(r, 2).value).strip()
        print("Excel task:", excel_task)

        if excel_task == task_name.strip():

            print("FOUND TASK:", excel_task)

            for c in range(3, open_col + 1):

                cell = ws.cell(r, c)
                color = str(cell.fill.start_color.rgb)

                if color in status_colors:
                    print("Employee:", headers[c - 1], "Color:", color)
                    employees.append(headers[c - 1])

            break

    print("Employees found:", employees)

    emails = []
    for emp in employees:
        email = get_employee_email(emp)

        if email:
            emails.append(email)
            print(emp, "=>", email)

        else:
            print("EMAIL NOT FOUND:", emp)

    return {
        "task": task_name,
        "employees": employees,
        "emails": emails
    }
    
@app.route("/")
def home():
    return app.send_static_file("index.html")

# ---------------------------
# 5. Upload endpoint
# ---------------------------
@app.route("/upload", methods=["POST"])


def upload():
    global LAST_UPLOADED_FILE
    global LAST_DRIVE_FILE_ID
    print("PID:", os.getpid())
    print("DB SIZE:", len(DB))
    print("DRIVE:", LAST_DRIVE_FILE_ID)
    

    try:
        DB.clear()

        files = request.files.getlist("files")
        results = []

        os.makedirs("uploads", exist_ok=True)

        for file in files:

            unique_name = f"{uuid.uuid4()}_{file.filename}"
            filepath = os.path.join("uploads", unique_name)

            file.save(filepath)

            print("FILE:", filepath)
            print("SIZE:", os.path.getsize(filepath))

            drive_file_id = upload_file(filepath)

            print("Drive ID:", drive_file_id)

            LAST_UPLOADED_FILE = filepath
            LAST_DRIVE_FILE_ID = drive_file_id
            print("LAST_DRIVE_FILE_ID =", LAST_DRIVE_FILE_ID)

            tasks = extract_tasks(filepath)
            print("Tasks extracted:", len(tasks))

            DB.append({
                "file": file.filename,
                "path": filepath,
                "tasks": tasks
            })

            results.append({
                "file": file.filename,
                "tasks_extracted": len(tasks)
            })


        return jsonify({
            "message": "Uploaded + normalized",
            "files": len(files),
            "results": results
        })


    except Exception as e:

        print("UPLOAD ERROR:")
        traceback.print_exc()

        return jsonify({
            "error": str(e)
        }), 500
    
@app.route("/debug-db")
def debug_db():
    print("BEFORE get_employee_sheet")
    sheet = get_employee_sheet()
    print("AFTER get_employee_sheet")
    print("BEFORE get_all_records")

    rows = sheet.get_all_records()
    print("AFTER get_all_records")

    return jsonify({
        "employees": rows,
        "count": len(rows)
    })

def create_task_logic(
    task,
    employees,
    open_date,
    close_date,
    drive_file_id,
    emails=None
):
    emails = emails or []

    print("===== CREATE TASK LOGIC =====")
    print("Task:", task)
    print("Employees:", employees)
    print("Open:", open_date)
    print("Close:", close_date)

    if not emails:
        print("No emails received, looking up...")

        emails = []

        for employee in employees:
            email = get_employee_email(employee)

            if email:
                emails.append(email)
            else:
                print("Missing email for:", employee)

    # -----------------------------
    # Download Excel
    # -----------------------------

    local_file = download_file(drive_file_id)

    try:

        wb = load_workbook(local_file)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]

        print("Headers:", headers)

        # -----------------------------
        # Find employee columns
        # -----------------------------

        employee_cols = []

        for employee in employees:

            found = False

            for i, header in enumerate(headers):

                if (
                    header
                    and str(header).strip().lower()
                    == employee.strip().lower()
                ):
                    employee_cols.append(i + 1)
                    found = True
                    break

            if not found:
                return {
                    "success": False,
                    "message": f"Employee '{employee}' not found"
                }

        # -----------------------------
        # Find empty row
        # -----------------------------

        new_row = ws.max_row + 1

        for r in range(2, ws.max_row + 1):

            if ws.cell(r, 2).value is None:
                new_row = r
                break

        # -----------------------------
        # Columns
        # -----------------------------

        task_col = 2

        open_col = headers.index("Open") + 1
        close_col = headers.index("Close") + 1

        # -----------------------------
        # Insert task
        # -----------------------------

        ws.cell(new_row, task_col).value = task

        # -----------------------------
        # Dates
        # -----------------------------

        open_dt = None
        close_dt = None

        if open_date:
            open_dt = datetime.strptime(
                open_date,
                "%Y-%m-%d"
            )

        if close_date:
            close_dt = datetime.strptime(
                close_date,
                "%Y-%m-%d"
            )

        if open_dt:

            ws.cell(
                new_row,
                open_col
            ).value = open_dt

            ws.cell(
                new_row,
                open_col
            ).number_format = "dd-mmm"

        if close_dt:

            ws.cell(
                new_row,
                close_col
            ).value = close_dt

            ws.cell(
                new_row,
                close_col
            ).number_format = "dd-mmm"

        # -----------------------------
        # Assign employees
        # -----------------------------

        for col in employee_cols:

            cell = ws.cell(
                new_row,
                col
            )

            cell.value = 1

            cell.fill = PatternFill(
                fill_type="solid",
                start_color="FF7A3F00",
                end_color="FF7A3F00"
            )

        # -----------------------------
        # Save
        # -----------------------------

        wb.save(local_file)

        print("Workbook saved successfully.")

        # -----------------------------
        # Upload to Drive
        # -----------------------------

        update_file(
            drive_file_id,
            local_file
        )

        print("Cloud Storage upload finished.")

        # -----------------------------
        # Refresh chatbot memory
        # -----------------------------

        new_tasks = extract_tasks(local_file)

        if DB:
            DB[0]["tasks"] = new_tasks

        print("CREATE SUCCESS")

        return {
            "success": True,
            "message": "CREATE SUCCESS",
            "task": task,
            "employees": employees,
            "emails": emails,
            "open": open_date,
            "close": close_date
        }

    finally:

        try:
            os.remove(local_file)
        except Exception as e:
            print("Could not remove temporary file:", e)

@app.route("/create-task", methods=["POST"])
def create_task():

    try:

        data = request.json

        print("===== CREATE TASK ENDPOINT =====")
        print(data)

        result = create_task_logic(
            task=data["task"],
            employees=data["employees"],
            open_date=data["open"],
            close_date=data["close"],
            drive_file_id=data["drive_file_id"],
            emails=data.get("emails", [])
        )

        if not result["success"]:
            return jsonify(result), 400

        return jsonify(result), 200

    except Exception as e:

        print("========== CREATE TASK ERROR ==========")

        import traceback
        traceback.print_exc()

        return jsonify({
            "success": False,
            "error": str(e)
        }), 500
    
# update excel
from openpyxl import load_workbook

@app.route("/update-excel", methods=["POST"])
def update_excel():



    data = request.json

    print("Received:", data)

    task = data["task"]
    employees = data["employees"]
    open_date = data["open"]
    close_date = data["close"]
    drive_file_id = data["drive_file_id"]
    local_file = download_file(drive_file_id)
    wb = load_workbook(local_file)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]

    print(headers)

    employee_cols = []
    
    for employee in employees:
        found = False
        
        for i, header in enumerate(headers):
            if header is not None and employee.lower() in str(header).lower():
                employee_cols.append(i + 1)
                found = True
                break
            
        if not found:
            return jsonify({
                "error": f"Employee '{employee}' not found."
                }), 400
    
    task_col = 2
    open_col = headers.index("Open") + 1
    close_col = headers.index("Close") + 1

    new_row = None
    
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 2).value is None:
            new_row = r
            break
        
    if new_row is None:
        new_row = ws.max_row + 1

    ws.cell(new_row, task_col).value = task

    if open_date:
        open_dt = datetime.strptime(open_date, "%Y-%m-%d")
        ws.cell(new_row, open_col).value = open_dt
        ws.cell(new_row, open_col).number_format = "dd-mmm"


    if close_date:
        close_dt = datetime.strptime(close_date, "%Y-%m-%d")
        ws.cell(new_row, close_col).value = close_dt
        ws.cell(new_row, close_col).number_format = "dd-mmm"
    

    for col in employee_cols:
        cell = ws.cell(new_row, col)
        
        cell.value = 1

        cell.fill = PatternFill(
            fill_type="solid",
            start_color="FF7A3F00",
            end_color="FF7A3F00")

    wb.save(local_file)
    update_file(drive_file_id, local_file)

    
    new_tasks = extract_tasks(local_file)
    
    print("New tasks extracted:", len(new_tasks))
    for t in new_tasks[-5:]:
        print(t)
    
    print("Downloaded file:", local_file)

    if DB:
        DB[0]["tasks"] = new_tasks

        print("Excel updated!")

        try:
            os.remove(local_file)
        except Exception as e:
            print("Couldn't delete temporary file:", e)

        return jsonify({
            "success": True,
            "task": task,
            "employees": employees,
            "emails": data.get("emails", []),
            "open": open_date,
            "close": close_date
            })


# ---------------------------
# 6. Dashboard API
# ---------------------------

@app.route("/dashboard", methods=["GET"])
def dashboard():

    global DB

    employee_counts = {}
    employee_status = {}
    unique_tasks = set()
    status_counts = {}
    status_colors = {}

    try:

        # ALWAYS reload latest Excel
        if LAST_DRIVE_FILE_ID:

            local_file = download_file(LAST_DRIVE_FILE_ID)

            tasks = extract_tasks(local_file)

            # rebuild cache
            DB = [{
                "file": "jobcard.xlsx",
                "path": local_file,
                "tasks": tasks
            }]

            print("Dashboard refreshed from Excel:", len(tasks))


        # Now calculate dashboard
        if DB:

            wb = load_workbook(DB[0]["path"])
            ws = wb.active

            headers = [cell.value for cell in ws[1]]

            open_col = headers.index("Open")

            for employee in headers[2:open_col]:

                if employee:
                    employee_counts.setdefault(employee, 0)
                    employee_status.setdefault(employee, {})


        for doc in DB:

            print(
                "Dashboard:",
                doc["file"],
                len(doc["tasks"])
            )

            for t in doc["tasks"]:

                unique_tasks.add(t["task"])

                emp = t["employee"]
                status = t["status"]

                status_colors[status] = t["color"]

                employee_counts[emp] = (
                    employee_counts.get(emp, 0) + 1
                )

                status_counts[status] = (
                    status_counts.get(status, 0) + 1
                )

                if emp not in employee_status:
                    employee_status[emp] = {}

                employee_status[emp][status] = (
                    employee_status[emp].get(status, 0) + 1
                )


        return jsonify({
            "total_tasks": len(unique_tasks),
            "total_employees": len(employee_counts),
            "tasks_per_employee": employee_counts,
            "status_breakdown": status_counts,
            "status_colors": status_colors,
            "employee_status": employee_status
        })


    except Exception as e:

        print("DASHBOARD ERROR:", e)

        return jsonify({
            "error": str(e)
        }), 500
    
from collections import defaultdict
@app.route("/chat", methods=["POST"])
def chat():
    try:
        print("ENTERED CHAT ROUTE")
        print("Step 1")
        # -------------------------
        # STEP 1 - Get employee message
        # -------------------------
        query = request.json.get("message", "")
        print("STEP 2", query)

        # -------------------------
        # STEP 2 - Let GPT determine
        # what action the employee wants
        # -------------------------
        command = client.chat.completions.create(
            model="gpt-4.1",
            response_format={"type": "json_object"},
            messages=[
                {
                    "role": "system",
                    "content": f"""
You are a command detector.

Always respond with a valid JSON object.
Do not return markdown.
Do not return explanations.
Your response must always be JSON.

DATE RULES:
- The current calendar year is {current_year}.
- If the user provides only a month and day without a year, always use {current_year}.
- If the user explicitly provides a year, always use that year.
- Never assume a previous year unless the user explicitly mentions it.
- Convert dates to YYYY-MM-DD format.

If the user wants to CREATE a task return:

{{
    "action":"create",
    "task":"",
    "employees":[],
    "open":"",
    "close":""
}}

If the user wants to DELETE a task return:

{{
    "action":"delete",
    "task":""
}}

If the user wants to update ONE field:

{{
  "action":"update",
  "task":"LCM Testing",
  "employee":"Sneha",
  "updates":[
    {{
      "field":"status",
      "value":"Done"
    }}
  ]
}}

If the user wants to ADD an employee return:

{{
  "action":"add_employee",
  "employee":"",
  "email":""
}}

If the user wants to update an employee email, return:

{{
    "action":"update_employee",
    "name":"",
    "email":""
}}

If the user wants to delete an employee, return:

{{
    "action":"delete_employee",
    "name":""
}}

If the user wants to update MULTIPLE fields:
{{
  "action":"update",
  "task":"LCM Testing",
  "employee":"",
  "updates":[
    {{
      "field":"status",
      "value":"Done"
    }},
    {{
      "field":"closeDate",
      "value":"2026-08-30"
    }}
  ]
}}

For status updates:
- Always include the employee whose status should change.
- If no employee is mentioned, set employee to "".

Otherwise return:

{{
    "action":"chat"
}}
"""
                },
                {
                    "role": "user",
                    "content": query
                }
            ]
        )
        print("STEP 3")

        command_json = json.loads(
            command.choices[0].message.content
        )
        print("STEP 4", command_json)



        action = command_json.get("action", "chat")
        print("STEP 5", action)

        # -------------------------
        # Make sure a Job Card exists
        # -------------------------
        if action != "chat":
            if len(DB) == 0:
                return jsonify({
                    "answer": "Please upload a Job Card first."
                })
        print("REACHED ACTION HANDLER:", action)

       

        # ====================================================
        # CREATE TASK
        # ====================================================

        if action == "create":
            print("CREATE BLOCK ENTERED")
            task = command_json.get("task", "")
            employees = command_json.get("employees", [])
            open_date = command_json.get("open", "")
            close_date = command_json.get("close", "")

            try:
                result = create_task_logic(
                    task=task,
                    employees=employees,
                    open_date=open_date,
                    close_date=close_date,
                    drive_file_id=LAST_DRIVE_FILE_ID
                )
                print("CREATE LOGIC RESULT:", result)

                if not result.get("success", False):
                    return jsonify({
                        "answer": result.get(
                            "message",
                            "Task creation failed."
                        ),
                        "success": False
                    }), 400
                
                
                try:
                    n8n_result = send_to_n8n(
                        action="create",
                        task=task,
                        employees=employees,
                        open_date=open_date,
                        close_date=close_date,
                        drive_file_id=LAST_DRIVE_FILE_ID
                    )
                    print("N8N EMAIL RESULT:", n8n_result)
                except Exception as n8n_error:
                    print("N8N EMAIL FAILED:", str(n8n_error))

                    import traceback
                    traceback.print_exc()
                    n8n_result = False

                return jsonify({
                    "answer": f"Task '{task}' created successfully.",
                    "success": True,
                    "email_sent": n8n_result
                }), 200

            except Exception as e:
                print("========== CHAT CREATE ERROR ==========")
                print("ERROR:", str(e))

                import traceback
                traceback.print_exc()

                return jsonify({
                    "answer": "Task creation failed.",
                    "success": False,
                    "error": str(e)
                }), 500

        # ====================================================
        # DELETE TASK
        # ====================================================

        elif action == "delete":
            task = command_json.get("task", "")

            try:
                result = delete_task_logic(
                    task=task,
                    drive_file_id=LAST_DRIVE_FILE_ID
                )
                print("DELETE LOGIC RESULT:", result)

                if not result.get("success", False):
                    return jsonify({
                        "answer": result.get(
                            "message",
                            "Task deletion failed."
                        ),
                        "success": False
                    }), 400

                employees = result.get("employees", [])

                n8n_result = send_to_n8n(
                    action="delete",
                    task=task,
                    employees=employees,
                    open_date="",
                    close_date="",
                    drive_file_id=LAST_DRIVE_FILE_ID
                )

                print("DELETE N8N RESULT:", n8n_result)
                return jsonify({
                    "answer": f"Task '{task}' deleted successfully.",
                    "success": True,
                    "email_sent": n8n_result
                }), 200

            except Exception as e:
                print("========== CHAT DELETE ERROR ==========")
                print("ERROR:", str(e))
                import traceback
                traceback.print_exc()

            return jsonify({
                "answer": "Task deletion failed.",
                "success": False,
                "error": str(e)
            }), 500


        # ====================================================
        # UPDATE TASK
        # ====================================================
        elif action == "update":
            print("========== CHAT UPDATE ==========")

            task = command_json.get("task", "")
            updates = command_json.get("updates", [])

            try:
                if command_json.get("employee"):
                    employees = [
                        command_json["employee"]
                    ]
                else:
                    local_file = download_file(
                    LAST_DRIVE_FILE_ID
                )

                details = find_task_details(
                    local_file,
                    task
                )
                os.remove(local_file)
                employees = details.get(
                    "employees",
                        []
                    )
                print("UPDATE EMPLOYEES:", employees)

                performed_updates = []
                for update in updates:
                    field = update["field"]
                    value = update["value"]
                    
                    result = update_task_logic(
                        task=task,
                        field=field,
                        value=value,
                        employees=employees,
                        selected_employee=command_json.get(
                            "employee",
                            ""
                        ),
                        drive_file_id=LAST_DRIVE_FILE_ID
                    )

                    print(
                        "UPDATE LOGIC RESULT:",
                        result
                    )
                    if not result.get("success", False):

                        return jsonify({
                            "answer": result.get(
                            "message",
                            "Task update failed."
                        ),
                        "success": False
                        }), 400

                    performed_updates.append({
                        "field": field,
                        "value": value
                    })

                    try:
                        n8n_result = send_to_n8n(
                        action="update",
                        task=task,
                        employees=employees,
                        open_date="",
                        close_date="",
                        drive_file_id=LAST_DRIVE_FILE_ID,
                        updates=performed_updates
                        )
                        print(
                            "UPDATE N8N RESULT:",
                            n8n_result
                        )

                    except Exception as n8n_error:
                        print(
                            "UPDATE N8N ERROR:",
                            str(n8n_error)
                        )
                        

                        import traceback
                        traceback.print_exc()   
                        n8n_result = False

                        return jsonify({
                            "answer": (
                                f"Task '{task}' updated successfully."
                            ),
                            "success": True,
                            "email_sent": n8n_result
                        }), 200

            except Exception as e:
                print("========== CHAT UPDATE ERROR ==========")
                print("ERROR:", str(e))

                import traceback
                traceback.print_exc()

                return jsonify({
                    "answer": "Task update failed.",
                    "success": False,
                    "error": str(e)
                }), 500
        # ====================================================
        # ADD EMPLOYEE
        # ====================================================
        elif action == "add_employee":
            employee = command_json.get("employee", "")
            email = command_json.get("email", "")

            try:
                result = add_employee_logic(
                employee=employee,
                email=email,
                drive_file_id=LAST_DRIVE_FILE_ID
                )
                print("ADD EMPLOYEE RESULT:", result)
                

                if not result.get("success", False):
                    return jsonify({
                        "answer": result.get(
                        "message",
                        "Failed to add employee."
                        ),
                        "success": False
                    }), 400

                try:
                    n8n_result = send_to_n8n(
                        action="welcome_employee",
                        task="",
                        employees=[employee],
                        open_date="",
                        close_date="",
                        drive_file_id=LAST_DRIVE_FILE_ID
                    )

                    print("WELCOME EMAIL N8N RESULT:", n8n_result)

                except Exception as n8n_error:
                    print("WELCOME EMAIL FAILED:", str(n8n_error))
                    import traceback
                    traceback.print_exc()
                    n8n_result = False

                    return jsonify({
                        "answer": f"{employee} added successfully.",
                        "success": True,
                        "email_sent": n8n_result
                    }), 200

            except Exception as e:
                print("========== ADD EMPLOYEE ERROR ==========")
                print("ERROR:", str(e))

                import traceback
                traceback.print_exc()
                return jsonify({
                    "answer": "Failed to add employee.",
                    "success": False,
                    "error": str(e)
                }), 500
               
        # ====================================================
        # NORMAL CHAT
        # ====================================================

        context = []
        emp_tasks = defaultdict(dict)

        for doc in DB:
            for t in doc["tasks"]:

                context.append(t)

                emp = t["employee"]
                task_name = t["task"]

                emp_tasks[emp][task_name] = {
                    "status": t["status"],
                    "open": t["open"],
                    "close": t["close"]
                }

        res = client.chat.completions.create(
            model="gpt-4.1",
            messages=[
                {
                    "role": "system",
                    "content": """
You are an AI Project Assistant.

You answer ONLY from the provided job card data.

RULES

1. Be concise.
2. Never invent information.
3. Use bullet points when appropriate.
4. If nothing matches say:
"No matching tasks were found."
"""
                },
                {
                    "role": "user",
                    "content": f"""
JOB CARD DATA

{json.dumps(context, indent=2)}

QUESTION

{query}
"""
                }
            ]
        )

        return jsonify({
            "answer": res.choices[0].message.content
        })
    except Exception as e:
        print("========== CHAT CRASH ==========")
        print("ERROR:", str(e))

    import traceback
    traceback.print_exc()

    return jsonify({
        "answer": "Backend crashed",
        "error": str(e)
    }), 500

def normalize_name(name):
    if not name:
        return ""

    return (
        name
        .strip()
        .lower()
        .replace(".", "")
        .replace("  ", " ")
    )
def get_employee_email(employee):

    sheet = get_employee_sheet()

    rows = sheet.get_all_records()

    employee = employee.strip().lower()

    for row in rows:

        sheet_name = str(row["employee_name"]).strip().lower()

        if employee == sheet_name:
            return row["email"]

    return None
            
def add_employee_sheet(name,email):

    sheet = get_employee_sheet()

    sheet.append_row([
        name,
        email
    ])

    print("Added to Google Sheet:", name)
def create_employee(employee, email, drive_file_id):
    global DB

    add_employee_sheet(employee, email)

    local_file = download_file(drive_file_id)

    wb = load_workbook(local_file)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]

    open_col = headers.index("Open") + 1

    ws.insert_cols(open_col)
    ws.cell(row=1, column=open_col).value = employee

    wb.save(local_file)

    update_file(drive_file_id, local_file)

    new_tasks = extract_tasks(local_file)

    DB = [{
        "file": "jobcard.xlsx",
        "path": local_file,
        "tasks": new_tasks
    }]

    os.remove(local_file)

    return True
def delete_task_logic(task, drive_file_id):

    local_file = None

    try:
        print("========== DELETE TASK LOGIC ==========")
        print("Task:", task)
        print("Drive File ID:", drive_file_id)

        # -----------------------------------------
        # Download latest Excel from Google Drive
        # -----------------------------------------

        local_file = download_file(drive_file_id)

        wb = load_workbook(local_file)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]

        print("Headers:", headers)

        # -----------------------------------------
        # Find task row
        # -----------------------------------------

        task_row = None

        for r in range(2, ws.max_row + 1):

            for c in range(1, ws.max_column + 1):

                cell_value = ws.cell(r, c).value

                if cell_value is None:
                    continue

                if str(cell_value).strip().lower() == task.strip().lower():

                    task_row = r
                    print(
                        f"Found task '{task}' "
                        f"at row {r}, column {c}"
                    )
                    break

            if task_row is not None:
                break

        # -----------------------------------------
        # Task not found
        # -----------------------------------------

        if task_row is None:

            print(f"TASK NOT FOUND: {task}")

            return {
                "success": False,
                "message": f"Task '{task}' not found."
            }

        # -----------------------------------------
        # Find Open column
        # -----------------------------------------

        if "Open" not in headers:

            return {
                "success": False,
                "message": "Open column not found in Excel."
            }

        open_col = headers.index("Open") + 1

        # -----------------------------------------
        # Find employees assigned to task
        # BEFORE deleting row
        # -----------------------------------------

        employees = []

        for c in range(3, open_col):

            header = headers[c - 1]

            if header is None:
                continue

            cell = ws.cell(task_row, c)

            if cell.value is not None:

                employees.append(
                    str(header).strip()
                )

        print("Employees assigned:", employees)

        # -----------------------------------------
        # Delete task row
        # -----------------------------------------

        ws.delete_rows(task_row, 1)

        print(
            f"Deleted row {task_row} "
            f"for task '{task}'."
        )

        # -----------------------------------------
        # Save Excel
        # -----------------------------------------

        wb.save(local_file)

        print("Workbook saved after delete.")

        # -----------------------------------------
        # Upload updated Excel back to Drive
        # -----------------------------------------

        update_file(
            drive_file_id,
            local_file
        )

        print("Drive updated after delete.")

        # -----------------------------------------
        # Refresh in-memory DB
        # -----------------------------------------

        new_tasks = extract_tasks(local_file)

        if DB:
            DB[0]["tasks"] = new_tasks

        print("DELETE SUCCESS")

        return {
            "success": True,
            "task": task,
            "employees": employees
        }

    finally:

        if local_file:

            try:
                os.remove(local_file)

            except Exception:
                pass
            
@app.route("/delete-task", methods=["POST"])
def delete_task():
    try:
        data = request.json

        task = data["task"]
        drive_file_id = data["drive_file_id"]

        print("========== DELETE TASK ROUTE ==========")
        print("Task:", task)

        result = delete_task_logic(
            task=task,
            drive_file_id=drive_file_id
        )

        print("DELETE LOGIC RESULT:", result)

        if not result.get("success", False):
            return jsonify(result), 404

        return jsonify({
            "success": True,
            "message": f"{task} deleted successfully.",
            "task": task,
            "employees": result.get("employees", [])
        }), 200

    except Exception as e:

        print("========== DELETE TASK ERROR ==========")
        print("ERROR:", str(e))

        import traceback
        traceback.print_exc()

        return jsonify({
            "success": False,
            "message": "Delete task failed.",
            "error": str(e)
        }), 500

@app.route("/task-details", methods=["POST"])
def task_details():
    try:
        data = request.json
        print("TASK DETAILS:", data)

        drive_file_id = data["drive_file_id"]

        local_file = download_file(drive_file_id)

        result = find_task_details(local_file, data["task"])

        employee_names = result.get("employees", [])

        emails = [get_employee_email(emp) for emp in employee_names]

        result["emails"] = emails

        result["drive_file_id"] = drive_file_id

        os.remove(local_file)

        return jsonify(result)

    except Exception as e:
        print("TASK DETAILS ERROR:", e)
        return jsonify({"error": str(e)}), 500
    
def update_task_logic(
    task,
    field,
    value,
    employees,
    selected_employee,
    drive_file_id
):

    local_file = None

    try:

        print("========== UPDATE TASK LOGIC ==========")
        print("Task:", task)
        print("Field:", field)
        print("Value:", value)
        print("Employees:", employees)
        print("Selected employee:", selected_employee)

        field = field.lower()

        field_map = {
            "name": "task",
            "opendate": "open",
            "closedate": "close"
        }

        field = field_map.get(field, field)

        # -----------------------------------------
        # Download Excel
        # -----------------------------------------

        local_file = download_file(drive_file_id)

        wb = load_workbook(local_file)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]

        open_col = headers.index("Open") + 1
        close_col = headers.index("Close") + 1

        # -----------------------------------------
        # Find task row
        # -----------------------------------------

        task_row = None

        for r in range(2, ws.max_row + 1):

            cell_value = ws.cell(r, 2).value

            if cell_value is None:
                continue

            if str(cell_value).strip().lower() == task.strip().lower():

                task_row = r
                break

        if task_row is None:

            return {
                "success": False,
                "message": f"Task '{task}' not found."
            }

        # ==================================================
        # STATUS UPDATE
        # ==================================================

        if field == "status":

            status_colors = {
                "done": "FF6AA84F",
                "due": "FFB7B7B7",
                "half-done": "FF6FA8DC",
                "redo": "FFFF0000",
                "late": "FFE84499",
                "on hold": "FFBF8E00",
                "almost ready": "FF00FFFF",
                "just started": "FF7A3F00",
                "na": "FFFFF2CC"
            }

            status = value.lower()

            if status not in status_colors:

                return {
                    "success": False,
                    "message": f"Unknown status: {value}"
                }

            fill = PatternFill(
                fill_type="solid",
                start_color=status_colors[status],
                end_color=status_colors[status]
            )

            # -----------------------------------------
            # Specific employee
            # -----------------------------------------

            if selected_employee:

                employee_col = None

                for i, header in enumerate(headers):

                    if header is None:
                        continue

                    if selected_employee.strip().lower() in str(header).strip().lower():

                        employee_col = i + 1

                        print(
                            "Matched employee column:",
                            employee_col
                        )

                        break

                if employee_col is None:

                    return {
                        "success": False,
                        "message": f"Employee '{selected_employee}' not found."
                    }

                cell = ws.cell(task_row, employee_col)

                if cell.value is not None:

                    print(
                        "Updating status:",
                        task_row,
                        employee_col
                    )

                    cell.fill = fill

            # -----------------------------------------
            # Everyone
            # -----------------------------------------

            else:

                for c in range(3, open_col):

                    cell = ws.cell(task_row, c)

                    if cell.value is not None:

                        cell.fill = fill

        # ==================================================
        # TASK NAME
        # ==================================================

        elif field == "task":

            ws.cell(task_row, 2).value = value

        # ==================================================
        # OPEN DATE
        # ==================================================

        elif field == "open":

            dt = datetime.strptime(value, "%Y-%m-%d")

            ws.cell(task_row, open_col).value = dt
            ws.cell(task_row, open_col).number_format = "dd-mmm"

        # ==================================================
        # CLOSE DATE
        # ==================================================

        elif field == "close":

            dt = datetime.strptime(value, "%Y-%m-%d")

            ws.cell(task_row, close_col).value = dt
            ws.cell(task_row, close_col).number_format = "dd-mmm"

        else:

            return {
                "success": False,
                "message": f"Unknown field: {field}"
            }

        # -----------------------------------------
        # Save Excel
        # -----------------------------------------

        wb.save(local_file)

        print("Workbook saved after update.")

        # -----------------------------------------
        # Upload to Drive
        # -----------------------------------------

        update_file(
            drive_file_id,
            local_file
        )

        print("Drive updated after update.")

        # -----------------------------------------
        # Refresh DB
        # -----------------------------------------

        global DB

        new_tasks = extract_tasks(local_file)

        DB = [{
            "file": "jobcard.xlsx",
            "path": local_file,
            "tasks": new_tasks
        }]

        print("UPDATE SUCCESS")

        return {
            "success": True,
            "task": task,
            "field": field,
            "value": value,
            "employees": employees,
            "selected_employee": selected_employee
        }

    finally:

        if local_file:

            try:
                os.remove(local_file)
            except Exception as e:
                print(
                    "Couldn't delete temporary file:",
                    e
                )

@app.route("/update-task", methods=["POST"])
def update_task():

    try:

        data = request.json

        print("========== UPDATE TASK ROUTE ==========")
        print(data)

        task = data["task"]
        field = data["field"]
        value = data["value"]

        employees = data.get("employees", [])

        selected_employee = data.get(
            "selected_employee",
            ""
        )

        drive_file_id = data["drive_file_id"]

        result = update_task_logic(
            task=task,
            field=field,
            value=value,
            employees=employees,
            selected_employee=selected_employee,
            drive_file_id=drive_file_id
        )

        print("UPDATE LOGIC RESULT:", result)

        if not result.get("success", False):

            return jsonify(result), 400

        return jsonify(result), 200

    except Exception as e:

        print("========== UPDATE TASK ERROR ==========")
        print("ERROR:", str(e))

        import traceback
        traceback.print_exc()

        return jsonify({
            "success": False,
            "message": "Update task failed.",
            "error": str(e)
        }), 500
def add_employee_logic(employee, email, drive_file_id):
    try:
        print("========== ADD EMPLOYEE LOGIC ==========")
        print("Employee:", employee)
        print("Email:", email)

        create_employee(
            employee,
            email,
            drive_file_id
        )

        print("EMPLOYEE CREATED:", employee)

        return {
            "success": True,
            "employee": employee,
            "email": email
        }

    except Exception as e:
        print("ADD EMPLOYEE LOGIC ERROR:", str(e))

        import traceback
        traceback.print_exc()

        return {
            "success": False,
            "message": str(e)
        }
    
@app.route("/add-employee", methods=["POST"])
def add_employee():

    try:
        data = request.json or {}

        employee = data.get("employee", "")
        email = data.get("email", "")
        drive_file_id = data.get("drive_file_id")

        print("========== ADD EMPLOYEE ROUTE ==========")
        print(data)

        result = add_employee_logic(
            employee=employee,
            email=email,
            drive_file_id=drive_file_id
        )

        print("ADD EMPLOYEE RESULT:", result)

        if not result.get("success"):
            return jsonify(result), 400

        # Send welcome email through n8n
        n8n_result = send_to_n8n(
            action="welcome_employee",
            task="",
            employees=[employee],
            open_date="",
            close_date="",
            drive_file_id=drive_file_id
        )

        print("WELCOME EMAIL N8N RESULT:", n8n_result)

        return jsonify({
            "success": True,
            "message": f"{employee} added successfully.",
            "employee": employee,
            "email": email,
            "email_sent": n8n_result
        }), 200

    except Exception as e:

        print("========== ADD EMPLOYEE ERROR ==========")
        print("ERROR:", str(e))

        import traceback
        traceback.print_exc()

        return jsonify({
            "success": False,
            "error": str(e)
        }), 500
@app.route("/download")
def download():

    global LAST_UPLOADED_FILE

    if not LAST_UPLOADED_FILE:
        return jsonify({
            "error": "No uploaded file."
        }), 404

    local_file = download_file(LAST_DRIVE_FILE_ID)

    return send_file(
        local_file,
        as_attachment=True,
        download_name="Updated_JobCard.xlsx"
    )

@app.route("/test-email/<name>")
def test_email(name):

    return jsonify({
        "name": name,
        "email": get_employee_email(name)
    })

@app.errorhandler(Exception)
def handle_error(e):
    import traceback
    traceback.print_exc()
    return jsonify({
        "error": str(e)
    }), 500

# ---------------------------
# RUN
# ---------------------------
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(
        host="0.0.0.0",
        port=port,
        debug=False
    )